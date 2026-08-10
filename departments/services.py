from django.db import transaction
from django.utils import timezone

from .models import Department
import csv

from openpyxl import Workbook
from django.http import HttpResponse

# ==========================================================
# CREATE DEPARTMENT
# ==========================================================

@transaction.atomic
def create_department(form, user):

    department = form.save(commit=False)

    department.created_by = user
    department.updated_by = user

    department.save()

    return department


# ==========================================================
# UPDATE DEPARTMENT
# ==========================================================

@transaction.atomic
def update_department(form, user):

    department = form.save(commit=False)

    department.updated_by = user

    department.save()

    return department


# ==========================================================
# DELETE DEPARTMENT
# ==========================================================

@transaction.atomic
def delete_department(department):

    department.delete()

    return True

# ==========================================================
# ACTIVATE DEPARTMENT
# ==========================================================

@transaction.atomic
def activate_department(department, user):

    department.status = "ACTIVE"
    department.updated_by = user
    department.updated_at = timezone.now()

    department.save(
        update_fields=[
            "status",
            "updated_by",
            "updated_at",
        ]
    )

    return department


# ==========================================================
# DEACTIVATE DEPARTMENT
# ==========================================================

@transaction.atomic
def deactivate_department(department, user):

    department.status = "INACTIVE"
    department.updated_by = user
    department.updated_at = timezone.now()

    department.save(
        update_fields=[
            "status",
            "updated_by",
            "updated_at",
        ]
    )

    return department


# ==========================================================
# TOGGLE DEPARTMENT STATUS
# ==========================================================

@transaction.atomic
def toggle_department_status(department, user):

    if department.status == "ACTIVE":
        department.status = "INACTIVE"
    else:
        department.status = "ACTIVE"

    department.updated_by = user
    department.updated_at = timezone.now()

    department.save(
        update_fields=[
            "status",
            "updated_by",
            "updated_at",
        ]
    )

    return department

import openpyxl

from django.db import transaction

from .models import Department


# ==========================================================
# BULK IMPORT DEPARTMENTS
# ==========================================================

@transaction.atomic
def bulk_import_departments(excel_file, user):

    workbook = openpyxl.load_workbook(excel_file)
    worksheet = workbook.active

    created_count = 0
    updated_count = 0
    skipped_count = 0

    error_rows = []

    # Skip Header Row
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2
    ):

        code = str(row[0]).strip() if row[0] else ""
        name = str(row[1]).strip() if row[1] else ""
        short_name = str(row[2]).strip() if row[2] else ""
        description = str(row[3]).strip() if row[3] else ""
        established_year = row[4]
        office_phone = str(row[5]).strip() if row[5] else ""
        office_email = str(row[6]).strip() if row[6] else ""
        building = str(row[7]).strip() if row[7] else ""
        total_faculty = row[8] or 0
        total_students = row[9] or 0
        status = str(row[10]).strip().upper() if row[10] else "ACTIVE"

        # -----------------------------
        # Required Validation
        # -----------------------------

        if not code or not name:

            skipped_count += 1

            error_rows.append({
                "row": row_number,
                "code": code,
                "name": name,
                "error": "Department Code and Name are required."
            })

            continue

        # -----------------------------
        # Status Validation
        # -----------------------------

        if status not in ["ACTIVE", "INACTIVE"]:

            status = "ACTIVE"

        # -----------------------------
        # Duplicate Name Check
        # -----------------------------

        if Department.objects.filter(name__iexact=name).exists():

            skipped_count += 1

            error_rows.append({
                "row": row_number,
                "code": code,
                "name": name,
                "error": "Department name already exists."
            })

            continue

        # -----------------------------
        # Duplicate Code Check
        # -----------------------------

        department = Department.objects.filter(
            code__iexact=code
        ).first()

                # -----------------------------
        # UPDATE EXISTING DEPARTMENT
        # -----------------------------

        if department:

            department.name = name
            department.short_name = short_name
            department.description = description
            department.established_year = established_year
            department.office_phone = office_phone
            department.office_email = office_email
            department.building = building
            department.total_faculty = total_faculty
            department.total_students = total_students
            department.status = status

            department.is_imported = True
            department.updated_by = user

            department.save()

            updated_count += 1

        # -----------------------------
        # CREATE NEW DEPARTMENT
        # -----------------------------

        else:

            Department.objects.create(

                code=code,
                name=name,
                short_name=short_name,
                description=description,
                established_year=established_year,

                office_phone=office_phone,
                office_email=office_email,
                building=building,

                total_faculty=total_faculty,
                total_students=total_students,

                status=status,

                is_imported=True,

                created_by=user,
                updated_by=user,
            )

            created_count += 1

    # -----------------------------
    # IMPORT SUMMARY
    # -----------------------------

    return {
        "created": created_count,
        "updated": updated_count,
        "skipped": skipped_count,
        "errors": error_rows,
    }
# ==========================================================
# EXPORT DEPARTMENTS TO EXCEL
# ==========================================================

def export_departments_excel():

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Departments"

    worksheet.append([
        "Department Code",
        "Department Name",
        "Short Name",
        "Description",
        "Established Year",
        "HOD",
        "Office Phone",
        "Office Email",
        "Building",
        "Total Faculty",
        "Total Students",
        "Status",
    ])

    departments = Department.objects.select_related(
        "hod"
    ).order_by("name")

    for department in departments:

        worksheet.append([

            department.code,
            department.name,
            department.short_name,
            department.description,
            department.established_year,

            department.hod.full_name
            if department.hod else "",

            department.office_phone,
            department.office_email,
            department.building,

            department.total_faculty,
            department.total_students,

            department.status,

        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="departments.xlsx"'

    workbook.save(response)

    return response


# ==========================================================
# EXPORT DEPARTMENTS TO CSV
# ==========================================================

def export_departments_csv():

    response = HttpResponse(
        content_type="text/csv"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="departments.csv"'

    writer = csv.writer(response)

    writer.writerow([
        "Department Code",
        "Department Name",
        "Short Name",
        "Description",
        "Established Year",
        "HOD",
        "Office Phone",
        "Office Email",
        "Building",
        "Total Faculty",
        "Total Students",
        "Status",
    ])

    departments = Department.objects.select_related(
        "hod"
    ).order_by("name")

    for department in departments:

        writer.writerow([

            department.code,
            department.name,
            department.short_name,
            department.description,
            department.established_year,

            department.hod.full_name
            if department.hod else "",

            department.office_phone,
            department.office_email,
            department.building,

            department.total_faculty,
            department.total_students,

            department.status,

        ])

    return response

# ==========================================================
# DOWNLOAD IMPORT ERROR REPORT
# ==========================================================

def download_error_report(error_rows):

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Import Errors"

    worksheet.append([
        "Row Number",
        "Department Code",
        "Department Name",
        "Error",
    ])

    for error in error_rows:

        worksheet.append([
            error.get("row"),
            error.get("code"),
            error.get("name"),
            error.get("error"),
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="department_import_errors.xlsx"'

    workbook.save(response)

    return response


# ==========================================================
# SAMPLE EXCEL FORMAT
# ==========================================================

def download_sample_excel():

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Departments Sample"

    worksheet.append([
        "Department Code",
        "Department Name",
        "Short Name",
        "Description",
        "Established Year",
        "Office Phone",
        "Office Email",
        "Building",
        "Total Faculty",
        "Total Students",
        "Status",
    ])

    worksheet.append([
        "CSE",
        "Computer Science and Engineering",
        "CSE",
        "Computer Science Department",
        2015,
        "9876543210",
        "cse@college.edu",
        "Block A",
        25,
        480,
        "ACTIVE",
    ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="department_sample.xlsx"'

    workbook.save(response)

    return response