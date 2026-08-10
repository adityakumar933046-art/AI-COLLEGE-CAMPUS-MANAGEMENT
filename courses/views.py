from results.models import Result
from timetable.models import Timetable
from notes.models import Note
from assignments.models import Assignment, AssignmentSubmission
from attendance.models import Attendance, AttendanceSession
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse

import csv
import uuid

from openpyxl import Workbook, load_workbook

from students.models import StudentProfile
from .models import Course, CourseMaterial
from accounts.decorators import admin_required

from .forms import (
    CourseCreateForm,
    CourseUpdateForm,
    CourseMaterialForm,
    CourseImportForm,
)

# ==========================================
# COURSE LIST
# ==========================================

from departments.models import Department
from teachers.models import TeacherProfile


# ==========================================
# COURSE LIST (ROLE-AWARE & SECURE)
# ==========================================

@login_required
def course_list(request):
    if request.user.role == "STUDENT":
        return redirect("student_courses")

    courses = Course.objects.select_related(
        "department",
        "teacher",
        "teacher__user",
    ).all().order_by("semester", "name")

    if request.user.role == "TEACHER":
        courses = courses.filter(teacher__user=request.user)

    search = (request.GET.get("search") or "").strip()
    if search:
        courses = courses.filter(
            Q(code__icontains=search)
            | Q(name__icontains=search)
            | Q(department__name__icontains=search)
            | Q(teacher__user__first_name__icontains=search)
            | Q(teacher__user__last_name__icontains=search)
        )

    department = request.GET.get("department")
    if department and department.isdigit():
        courses = courses.filter(department_id=int(department))

    semester = request.GET.get("semester")
    if semester and semester.isdigit():
        courses = courses.filter(semester=int(semester))

    status = request.GET.get("status")
    if status and status in ["ACTIVE", "INACTIVE"]:
        courses = courses.filter(status=status)

    teacher_id = request.GET.get("teacher")
    if teacher_id and teacher_id.isdigit():
        courses = courses.filter(teacher_id=int(teacher_id))

    per_page = request.GET.get("per_page", 10)
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 10

    paginator = Paginator(courses, per_page)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    all_departments = Department.objects.all().order_by("name")
    all_teachers = TeacherProfile.objects.select_related("user").all().order_by("user__first_name")

    return render(
        request,
        "courses/course_list.html",
        {
            "courses": page_obj,
            "page_obj": page_obj,
            "search": search,
            "department": department or "",
            "semester": semester or "",
            "status": status or "",
            "teacher": teacher_id or "",
            "per_page": per_page,
            "departments": all_departments,
            "teachers": all_teachers,
            "title": "Courses Management" if request.user.role == "ADMIN" else "My Assigned Courses",
        },
    )

# ==========================================
# COURSE DETAIL
# ==========================================

@login_required
@login_required
def course_detail(request, pk):
    course = get_object_or_404(
        Course.objects.select_related("department", "teacher", "teacher__user"),
        pk=pk,
    )
    
    # Fetch connected course resources
    materials = CourseMaterial.objects.filter(course=course).order_by("-uploaded_at")
    assignments = Assignment.objects.filter(course=course).order_by("-created_at")
    notes = Note.objects.filter(course=course).order_by("-uploaded_at")
    timetables = Timetable.objects.filter(course=course).order_by("day", "start_time")
    attendance_sessions = AttendanceSession.objects.filter(course=course).order_by("-attendance_date")

    # Context data for role-specific insights
    student_stats = None
    if request.user.role == "STUDENT":
        try:
            student = request.user.student_profile
            total_sessions = Attendance.objects.filter(session__course=course, student=student).count()
            present_sessions = Attendance.objects.filter(session__course=course, student=student, status="PRESENT").count()
            att_pct = round((present_sessions / total_sessions * 100), 1) if total_sessions > 0 else 0.0
            
            # Submissions for this student
            submissions = AssignmentSubmission.objects.filter(assignment__course=course, student=student)
            submitted_ids = submissions.values_list("assignment_id", flat=True)
            
            # Results
            results = Result.objects.filter(course=course, student=student)
            
            student_stats = {
                "attendance_pct": att_pct,
                "total_sessions": total_sessions,
                "present_sessions": present_sessions,
                "submitted_assignment_ids": list(submitted_ids),
                "results": results,
            }
        except Exception:
            pass

    return render(
        request,
        "courses/course_detail.html",
        {
            "course": course,
            "materials": materials,
            "assignments": assignments,
            "notes": notes,
            "timetables": timetables,
            "attendance_sessions": attendance_sessions,
            "student_stats": student_stats,
            "title": f"{course.code} - {course.name}",
        },
    )

# ==========================================
# CREATE COURSE
# ==========================================

@login_required
@admin_required
def course_create(request):
    if request.method == "POST":
        form = CourseCreateForm(request.POST)
        if form.is_valid():
            course = form.save(commit=False)
            course.created_by = request.user
            course.updated_by = request.user
            course.save()
            messages.success(request, "Course created successfully.")
            return redirect("course_list")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{error}")
    else:
        form = CourseCreateForm()

    return render(
        request,
        "courses/course_create.html",
        {
            "form": form
        },
    )
# ==========================================
# UPDATE COURSE
# ==========================================

@login_required
@admin_required
def course_update(request, pk):

    course = get_object_or_404(
        Course,
        pk=pk
    )

    if request.method == "POST":

        form = CourseUpdateForm(
            request.POST,
            instance=course
        )

        if form.is_valid():

            course = form.save(commit=False)

            course.updated_by = request.user

            course.save()

            messages.success(
                request,
                "Course updated successfully."
            )

            return redirect(
                "course_detail",
                pk=course.pk
            )

    else:

        form = CourseUpdateForm(
            instance=course
        )

    return render(
        request,
        "courses/course_update.html",
        {
            "form": form,
            "course": course,
        },
    )


# ==========================================
# DELETE COURSE
# ==========================================

@login_required
@admin_required
def course_delete(request, pk):
    course = get_object_or_404(Course, pk=pk)

    has_records = (
        AttendanceSession.objects.filter(course=course).exists()
        or Assignment.objects.filter(course=course).exists()
        or Result.objects.filter(course=course).exists()
        or Note.objects.filter(course=course).exists()
        or Timetable.objects.filter(course=course).exists()
        or CourseMaterial.objects.filter(course=course).exists()
    )

    if request.method == "POST":
        if has_records:
            course.status = "INACTIVE"
            course.updated_by = request.user
            course.save()
            messages.warning(
                request,
                f"Course '{course.code} - {course.name}' has existing academic records. It was set to INACTIVE to preserve historical data."
            )
        else:
            course.delete()
            messages.success(request, f"Course '{course.code}' deleted successfully.")

        return redirect("course_list")

    return render(
        request,
        "courses/course_delete.html",
        {
            "course": course,
            "has_records": has_records,
            "title": f"Delete / Deactivate Course {course.code}",
        },
    )


# ==========================================
# ACTIVATE COURSE
# ==========================================

@login_required
@admin_required
def activate_course(request, pk):

    course = get_object_or_404(
        Course,
        pk=pk
    )

    course.status = "ACTIVE"
    course.updated_by = request.user
    course.save()

    messages.success(
        request,
        "Course activated successfully."
    )

    return redirect(
        "course_detail",
        pk=pk
    )


# ==========================================
# DEACTIVATE COURSE
# ==========================================

@login_required
@admin_required
def deactivate_course(request, pk):

    course = get_object_or_404(
        Course,
        pk=pk
    )

    course.status = "INACTIVE"
    course.updated_by = request.user
    course.save()

    messages.success(
        request,
        "Course deactivated successfully."
    )

    return redirect(
        "course_detail",
        pk=pk
    )


# ==========================================
# STUDENT COURSES
# ==========================================

@login_required
@login_required
def student_courses(request):
    try:
        student = request.user.student_profile
    except Exception:
        return redirect("dashboard_redirect")

    courses = Course.objects.filter(
        department=student.department,
        semester=student.semester,
        status="ACTIVE"
    ).select_related(
        "department",
        "teacher",
        "teacher__user"
    ).order_by("name")

    # Enrich courses with subject-wise attendance and assignment counts
    course_list_enriched = []
    for c in courses:
        total_att = Attendance.objects.filter(session__course=c, student=student).count()
        present_att = Attendance.objects.filter(session__course=c, student=student, status="PRESENT").count()
        att_pct = round((present_att / total_att * 100), 1) if total_att > 0 else 0.0
        
        ass_count = Assignment.objects.filter(course=c).count()
        notes_count = Note.objects.filter(course=c).count()

        course_list_enriched.append({
            "course": c,
            "attendance_pct": att_pct,
            "total_att": total_att,
            "assignments_count": ass_count,
            "notes_count": notes_count,
        })

    return render(
        request,
        "courses/student_courses.html",
        {
            "enriched_courses": course_list_enriched,
            "student": student,
            "title": "My Courses",
        },
    )

# ==========================================
# COURSE MATERIAL LIST
# ==========================================

@login_required
def course_materials(request, course_id):

    course = get_object_or_404(
        Course,
        pk=course_id
    )

    materials = CourseMaterial.objects.filter(
        course=course
    ).select_related(
        "uploaded_by"
    ).order_by(
        "-uploaded_at"
    )

    return render(
        request,
        "courses/course_materials.html",
        {
            "course": course,
            "materials": materials,
        },
    )


# ==========================================
# UPLOAD COURSE MATERIAL
# ==========================================

@login_required
@admin_required
def upload_course_material(request):

    if request.method == "POST":

        form = CourseMaterialForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():

            material = form.save(commit=False)

            material.uploaded_by = request.user

            material.save()

            messages.success(
                request,
                "Course material uploaded successfully."
            )

            return redirect(
                "course_materials",
                course_id=material.course.id
            )

    else:

        form = CourseMaterialForm()

    return render(
        request,
        "courses/upload_course_material.html",
        {
            "form": form,
        },
    )


# ==========================================
# DELETE COURSE MATERIAL
# ==========================================

@login_required
@admin_required
def delete_course_material(request, pk):

    material = get_object_or_404(
        CourseMaterial,
        pk=pk
    )

    course_id = material.course.id

    if request.method == "POST":

        material.delete()

        messages.success(
            request,
            "Course material deleted successfully."
        )

        return redirect(
            "course_materials",
            course_id=course_id
        )

    return render(
        request,
        "courses/delete_course_material.html",
        {
            "material": material,
        },
    )
# ==========================================
# IMPORT COURSES FROM EXCEL
# ==========================================

@login_required
@admin_required
def import_courses(request):

    if request.method == "POST":

        form = CourseImportForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():

            workbook = load_workbook(
                request.FILES["excel_file"]
            )

            sheet = workbook.active

            batch_id = str(uuid.uuid4())[:8]

            imported = 0
            skipped = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):

                if not row or not row[0]:
                    continue

                code = str(row[0]).strip()

                if Course.objects.filter(
                    code__iexact=code
                ).exists():

                    skipped += 1
                    continue

                Course.objects.create(

                    code=code,

                    name=row[1],

                    department_id=row[2],

                    teacher_id=row[3] if row[3] else None,

                    semester=row[4],

                    credits=row[5],

                    status=row[6] or "ACTIVE",

                    imported_from_excel=True,

                    import_batch=batch_id,

                    created_by=request.user,

                    updated_by=request.user,

                )

                imported += 1

            messages.success(
                request,
                f"{imported} courses imported successfully. {skipped} duplicate courses skipped."
            )

            return redirect("course_list")

    else:

        form = CourseImportForm()

    return render(
        request,
        "courses/import_courses.html",
        {
            "form": form
        }
    )


# ==========================================
# EXPORT COURSES TO EXCEL
# ==========================================

@login_required
@admin_required
def export_courses_excel(request):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Courses"

    sheet.append([
        "Code",
        "Course Name",
        "Department ID",
        "Teacher ID",
        "Semester",
        "Credits",
        "Status",
    ])

    for course in Course.objects.all():

        sheet.append([

            course.code,

            course.name,

            course.department_id,

            course.teacher_id,

            course.semester,

            course.credits,

            course.status,

        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="courses.xlsx"'

    workbook.save(response)

    return response


# ==========================================
# EXPORT COURSES TO CSV
# ==========================================

@login_required
@admin_required
def export_courses_csv(request):

    response = HttpResponse(
        content_type="text/csv"
    )

    response["Content-Disposition"] = 'attachment; filename="courses.csv"'

    writer = csv.writer(response)

    writer.writerow([
        "Code",
        "Course Name",
        "Department ID",
        "Teacher ID",
        "Semester",
        "Credits",
        "Status",
    ])

    for course in Course.objects.all():

        writer.writerow([

            course.code,

            course.name,

            course.department_id,

            course.teacher_id,

            course.semester,

            course.credits,

            course.status,

        ])

    return response


# ==========================================
# DOWNLOAD SAMPLE EXCEL
# ==========================================

@login_required
@admin_required
def download_course_sample(request):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Sample"

    sheet.append([
        "Code",
        "Course Name",
        "Department ID",
        "Teacher ID",
        "Semester",
        "Credits",
        "Status",
    ])

    sheet.append([
        "CS101",
        "Python Programming",
        1,
        1,
        1,
        4,
        "ACTIVE",
    ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="sample_courses.xlsx"'

    workbook.save(response)

    return response