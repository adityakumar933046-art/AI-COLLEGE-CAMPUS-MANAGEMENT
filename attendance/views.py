from courses.models import Course
import csv
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.http import HttpResponse

from accounts.decorators import admin_required, teacher_required, admin_or_teacher_required
from students.models import StudentProfile
from teachers.models import TeacherProfile
from timetable.models import Timetable
from .models import AttendanceSession, Attendance
from .forms import AttendanceSessionForm, AttendanceForm


# =====================================================
# ATTENDANCE DASHBOARD
# =====================================================
@login_required
def attendance_dashboard(request):
    total_sessions = AttendanceSession.objects.count()
    total_attendance = Attendance.objects.count()
    present_count = Attendance.objects.filter(status="PRESENT").count()
    absent_count = Attendance.objects.filter(status="ABSENT").count()
    
    pct = round((present_count / total_attendance) * 100, 1) if total_attendance > 0 else 0.0

    context = {
        "total_sessions": total_sessions,
        "total_attendance": total_attendance,
        "present_count": present_count,
        "absent_count": absent_count,
        "overall_percentage": pct,
    }
    return render(request, "attendance/dashboard.html", context)


# =====================================================
# ATTENDANCE SESSION LIST
# =====================================================
@login_required
def attendance_session_list(request):
    sessions = AttendanceSession.objects.select_related(
        "department", "course", "teacher", "teacher__user"
    ).all().order_by("-attendance_date")

    if request.user.role == "TEACHER":
        sessions = sessions.filter(teacher__user=request.user)

    search = request.GET.get("search")
    if search:
        sessions = sessions.filter(
            Q(course__name__icontains=search) | Q(course__code__icontains=search)
        )

    semester = request.GET.get("semester")
    if semester:
        sessions = sessions.filter(semester=semester)

    section = request.GET.get("section")
    if section:
        sessions = sessions.filter(section__iexact=section)

    paginator = Paginator(sessions, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "sessions": page_obj,
        "page_obj": page_obj,
        "search": search,
        "semester": semester,
        "section": section,
    }
    return render(request, "attendance/session_list.html", context)


# =====================================================
# CREATE ATTENDANCE SESSION
# =====================================================
@login_required
@teacher_required
def create_attendance_session(request):
    if request.method == "POST":
        form = AttendanceSessionForm(request.POST)
        if form.is_valid():
            session = form.save(commit=False)
            session.created_by = request.user
            session.save()
            messages.success(request, "Attendance session created successfully.")
            return redirect("mark_attendance", session_id=session.id)
    else:
        form = AttendanceSessionForm()

    return render(request, "attendance/create_session.html", {"form": form})


# =====================================================
# TAKE ATTENDANCE FROM TIMETABLE
# =====================================================
@login_required
@teacher_required
def take_attendance_from_timetable(request, timetable_id):
    teacher = get_object_or_404(TeacherProfile, user=request.user)
    timetable = get_object_or_404(Timetable, id=timetable_id, teacher=teacher)
    
    now = timezone.localtime()
    class_start = timezone.make_aware(datetime.combine(date.today(), timetable.start_time))
    class_end = timezone.make_aware(datetime.combine(date.today(), timetable.end_time))
    grace_end = class_end + timedelta(minutes=30)

    if now < class_start:
        messages.warning(request, "Class has not started yet.")
        return redirect("teacher_timetable")

    if now > grace_end:
        messages.error(request, "Attendance window has closed.")
        return redirect("teacher_timetable")

    session, created = AttendanceSession.objects.get_or_create(
        department=timetable.department,
        course=timetable.course,
        teacher=teacher,
        semester=timetable.semester,
        section=timetable.section,
        attendance_date=date.today(),
        defaults={
            "created_by": request.user,
            "status": "OPEN",
            "start_time": timetable.start_time,
            "end_time": timetable.end_time,
        }
    )
    return redirect("mark_attendance", session_id=session.id)


# =====================================================
# MARK ATTENDANCE
# =====================================================
@login_required
@teacher_required
def mark_attendance(request, session_id):
    session = get_object_or_404(AttendanceSession, id=session_id)
    students = StudentProfile.objects.filter(
        department=session.department,
        semester=session.semester
    )

    attendance_records = Attendance.objects.filter(session=session)
    marked_dict = {record.student_id: record.status for record in attendance_records}

    context = {
        "session": session,
        "students": students,
        "marked_dict": marked_dict,
    }
    return render(request, "attendance/mark_attendance.html", context)


# =====================================================
# SAVE ATTENDANCE
# =====================================================
@login_required
@teacher_required
@login_required
@admin_or_teacher_required
def save_attendance(request, session_id):
    session = get_object_or_404(AttendanceSession, id=session_id)
    
    if request.user.role == "TEACHER" and hasattr(request.user, "teacher_profile"):
        if session.teacher != request.user.teacher_profile:
            messages.error(request, "You are not authorized to modify attendance for this session.")
            return redirect("attendance_session_list")

    if request.method == "POST":
        saved_count = 0
        for key, value in request.POST.items():
            if key.startswith("status_"):
                student_id = key.split("_")[1]
                try:
                    student = StudentProfile.objects.get(id=student_id)
                    remarks = request.POST.get(f"remarks_{student_id}", "").strip()
                    Attendance.objects.update_or_create(
                        session=session,
                        student=student,
                        defaults={
                            "status": value,
                            "remarks": remarks,
                            "marked_by": request.user,
                        }
                    )
                    saved_count += 1
                except StudentProfile.DoesNotExist:
                    continue

        messages.success(request, f"Attendance saved successfully for {saved_count} student(s).")
        return redirect("attendance_session_list")

    return redirect("mark_attendance", session_id=session.id)


# =====================================================
# ATTENDANCE HISTORY & DETAILS
# =====================================================
@login_required
def attendance_history(request):
    return redirect("attendance_session_list")

@login_required
def attendance_detail(request, pk):
    session = get_object_or_404(AttendanceSession, pk=pk)
    records = Attendance.objects.filter(session=session).select_related("student", "student__user")
    return render(request, "attendance/attendance_detail.html", {"session": session, "records": records})

@login_required
@teacher_required
def attendance_update(request, pk):
    return redirect("mark_attendance", session_id=pk)

@login_required
@admin_required
def attendance_delete(request, pk):
    session = get_object_or_404(AttendanceSession, pk=pk)
    session.delete()
    messages.success(request, "Attendance session deleted.")
    return redirect("attendance_session_list")


# =====================================================
# STUDENT ATTENDANCE SUMMARY
# =====================================================
@login_required
# =====================================================
# STUDENT ATTENDANCE SUMMARY (ADVANCED)
# =====================================================
@login_required
def student_attendance(request):
    try:
        student = StudentProfile.objects.select_related("department", "user").get(user=request.user)
    except StudentProfile.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect("home")

    # Academic courses for student
    academic_courses = Course.objects.filter(
        department=student.department,
        semester=student.semester
    ).select_related("teacher", "teacher__user")

    # Base Attendance queryset for student
    attendance_qs = Attendance.objects.filter(
        student=student
    ).select_related(
        "session",
        "session__course",
        "session__teacher",
        "session__teacher__user",
        "session__department"
    )

    # Filter parameters
    course_filter = request.GET.get("course", "").strip()
    from_date = request.GET.get("from_date", "").strip()
    to_date = request.GET.get("to_date", "").strip()
    status_filter = request.GET.get("status", "").strip()
    search_q = request.GET.get("search", "").strip()

    if course_filter and course_filter.isdigit():
        attendance_qs = attendance_qs.filter(session__course_id=int(course_filter))

    if from_date:
        attendance_qs = attendance_qs.filter(session__attendance_date__gte=from_date)

    if to_date:
        attendance_qs = attendance_qs.filter(session__attendance_date__lte=to_date)

    if status_filter in ["PRESENT", "ABSENT", "LATE", "MEDICAL"]:
        attendance_qs = attendance_qs.filter(status=status_filter)

    if search_q:
        attendance_qs = attendance_qs.filter(
            Q(session__course__name__icontains=search_q) |
            Q(session__course__code__icontains=search_q) |
            Q(session__remarks__icontains=search_q) |
            Q(remarks__icontains=search_q)
        )

    # Order class history by date descending
    class_history = attendance_qs.order_by("-session__attendance_date", "-session__lecture_no")

    # Calculate overall KPIs
    total_classes = attendance_qs.count()
    total_present = attendance_qs.filter(status="PRESENT").count()
    total_absent = attendance_qs.filter(status="ABSENT").count()
    total_late = attendance_qs.filter(status="LATE").count()
    total_medical = attendance_qs.filter(status="MEDICAL").count()

    overall_pct = round((total_present / total_classes) * 100, 1) if total_classes > 0 else 0.0

    # Course-wise breakdown (from student's academic courses)
    course_attendance = []
    low_attendance_courses = []
    
    # Chart data containers
    chart_labels = []
    chart_percentages = []
    chart_present = []
    chart_absent = []

    for crs in academic_courses:
        c_recs = Attendance.objects.filter(student=student, session__course=crs)
        c_total = c_recs.count()
        c_present = c_recs.filter(status="PRESENT").count()
        c_absent = c_recs.filter(status="ABSENT").count()
        c_pct = round((c_present / c_total) * 100, 1) if c_total > 0 else 0.0

        item = {
            "course": crs,
            "total": c_total,
            "present": c_present,
            "absent": c_absent,
            "percentage": c_pct,
            "is_low": c_pct < 75.0 and c_total > 0,
        }
        course_attendance.append(item)

        if item["is_low"]:
            low_attendance_courses.append(item)

        chart_labels.append(crs.code if crs.code else crs.name[:10])
        chart_percentages.append(c_pct)
        chart_present.append(c_present)
        chart_absent.append(c_absent)

    # Monthly Summary
    monthly_dict = {}
    for rec in attendance_qs:
        dt = rec.session.attendance_date
        if dt:
            month_key = dt.strftime("%B %Y")
            if month_key not in monthly_dict:
                monthly_dict[month_key] = {"total": 0, "present": 0, "absent": 0}
            monthly_dict[month_key]["total"] += 1
            if rec.status == "PRESENT":
                monthly_dict[month_key]["present"] += 1
            else:
                monthly_dict[month_key]["absent"] += 1

    monthly_summary = []
    for m_key, m_val in monthly_dict.items():
        m_pct = round((m_val["present"] / m_val["total"]) * 100, 1) if m_val["total"] > 0 else 0.0
        monthly_summary.append({
            "month": m_key,
            "total": m_val["total"],
            "present": m_val["present"],
            "absent": m_val["absent"],
            "percentage": m_pct,
        })

    context = {
        "student": student,
        "academic_courses": academic_courses,
        "overall_percentage": overall_pct,
        "total_classes": total_classes,
        "total_present": total_present,
        "total_absent": total_absent,
        "total_late": total_late,
        "total_medical": total_medical,
        "course_attendance": course_attendance,
        "low_attendance_courses": low_attendance_courses,
        "class_history": class_history,
        "monthly_summary": monthly_summary,
        "chart_labels": chart_labels,
        "chart_percentages": chart_percentages,
        "chart_present": chart_present,
        "chart_absent": chart_absent,
        "selected_course": course_filter,
        "selected_from_date": from_date,
        "selected_to_date": to_date,
        "selected_status": status_filter,
        "search_query": search_q,
    }
    return render(request, "attendance/student_attendance.html", context)


# =====================================================
# ATTENDANCE DEFAULTERS (< 75%)
# =====================================================
@login_required
@admin_or_teacher_required
def attendance_defaulters(request):
    students = StudentProfile.objects.select_related("department", "user").all()
    defaulters = []

    for s in students:
        recs = Attendance.objects.filter(student=s)
        tot = recs.count()
        if tot > 0:
            prs = recs.filter(status="PRESENT").count()
            pct = round((prs / tot) * 100, 1)
            if pct < 75.0:
                defaulters.append({
                    "student": s,
                    "total": tot,
                    "present": prs,
                    "percentage": pct,
                })

    return render(request, "attendance/defaulters.html", {"defaulters": defaulters})


# =====================================================
# ATTENDANCE EXPORTS (FILTER-AWARE & ROLE-SECURE)
# =====================================================
# ==========================================
# ATTENDANCE EXPORT EXCEL
# ==========================================
# ==========================================
# ATTENDANCE FILTER & EXPORT HELPERS
# ==========================================
def clean_query_param(val):
    if val is None:
        return None
    val_str = str(val).strip()
    if val_str.lower() in ("", "none", "null", "undefined"):
        return None
    return val_str

def get_filtered_attendance_records(request):
    user = request.user
    records = Attendance.objects.select_related(
        "student",
        "student__user",
        "student__department",
        "session",
        "session__course",
        "session__teacher",
        "session__teacher__user"
    ).all()

    if user.role == "STUDENT":
        records = records.filter(student__user=user)
    elif user.role == "TEACHER":
        records = records.filter(session__teacher__user=user)

    search = clean_query_param(request.GET.get("search"))
    if search:
        records = records.filter(
            Q(session__course__name__icontains=search) |
            Q(session__course__code__icontains=search) |
            Q(student__user__first_name__icontains=search) |
            Q(student__user__last_name__icontains=search) |
            Q(student__roll_no__icontains=search)
        )

    semester = clean_query_param(request.GET.get("semester"))
    if semester and semester.isdigit():
        records = records.filter(session__semester=int(semester))

    section = clean_query_param(request.GET.get("section"))
    if section:
        records = records.filter(session__section__iexact=section)

    return records.order_by("-session__attendance_date", "session__course__code")


# ==========================================
# ATTENDANCE EXPORT EXCEL
# ==========================================
@login_required
def export_attendance_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    records = get_filtered_attendance_records(request)

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Attendance Records"

    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")

    headers = ["Date", "Course Code", "Course Title", "Faculty Member", "Student Name", "Roll Number", "Semester", "Section", "Status"]
    ws1.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    student_summary = {}
    for rec in records:
        date_str = rec.session.attendance_date.strftime("%Y-%m-%d") if (rec.session and rec.session.attendance_date) else "N/A"
        teacher_name = rec.session.teacher.full_name if (rec.session and rec.session.teacher) else "N/A"
        student_name = rec.student.full_name if rec.student else "N/A"
        roll_no = rec.student.roll_no if rec.student else "N/A"

        ws1.append([
            date_str,
            rec.session.course.code if (rec.session and rec.session.course) else "N/A",
            rec.session.course.name if (rec.session and rec.session.course) else "N/A",
            teacher_name,
            student_name,
            roll_no,
            rec.session.semester if rec.session else "N/A",
            rec.session.section if rec.session else "N/A",
            rec.status
        ])

        if roll_no not in student_summary:
            student_summary[roll_no] = {"name": student_name, "total": 0, "present": 0, "absent": 0}
        student_summary[roll_no]["total"] += 1
        if rec.status == "PRESENT":
            student_summary[roll_no]["present"] += 1
        else:
            student_summary[roll_no]["absent"] += 1

    ws2 = wb.create_sheet(title="Summary Statistics")
    summary_headers = ["Roll Number", "Student Name", "Total Lectures", "Present", "Absent", "Attendance %"]
    ws2.append(summary_headers)

    for col_num in range(1, len(summary_headers) + 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for roll_no, stats in student_summary.items():
        pct = round((stats["present"] / stats["total"]) * 100, 1) if stats["total"] > 0 else 0
        ws2.append([roll_no, stats["name"], stats["total"], stats["present"], stats["absent"], f"{pct}%"])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="smart_campus_attendance.xlsx"'
    wb.save(response)
    return response


# ==========================================
# ATTENDANCE EXPORT CSV
# ==========================================
@login_required
def export_attendance_csv(request):
    import csv

    records = get_filtered_attendance_records(request)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="smart_campus_attendance.csv"'

    writer = csv.writer(response)
    writer.writerow(["Date", "Course Code", "Course Title", "Student Name", "Roll Number", "Semester", "Section", "Status"])

    for rec in records:
        writer.writerow([
            rec.session.attendance_date.strftime("%Y-%m-%d") if (rec.session and rec.session.attendance_date) else "N/A",
            rec.session.course.code if (rec.session and rec.session.course) else "N/A",
            rec.session.course.name if (rec.session and rec.session.course) else "N/A",
            rec.student.full_name if rec.student else "N/A",
            rec.student.roll_no if rec.student else "N/A",
            rec.session.semester if rec.session else "N/A",
            rec.session.section if rec.session else "N/A",
            rec.status
        ])

    return response
