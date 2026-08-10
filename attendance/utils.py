from datetime import date
import csv

from django.db.models import Count, Q
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font

from .models import Attendance


# ==========================================
# Attendance Percentage
# ==========================================

def calculate_percentage(present, total):
    if total == 0:
        return 0
    return round((present / total) * 100, 2)


# ==========================================
# Student Attendance Summary
# ==========================================

def student_summary(student):

    queryset = Attendance.objects.filter(student=student)

    total = queryset.count()

    present = queryset.filter(status="PRESENT").count()

    absent = queryset.filter(status="ABSENT").count()

    late = queryset.filter(status="LATE").count()

    medical = queryset.filter(status="MEDICAL").count()

    return {
        "total": total,
        "present": present,
        "absent": absent,
        "late": late,
        "medical": medical,
        "percentage": calculate_percentage(
            present,
            total
        ),
    }


# ==========================================
# Course Attendance
# ==========================================

def course_summary(course):

    queryset = Attendance.objects.filter(
        session__course=course
    )

    total = queryset.count()

    present = queryset.filter(
        status="PRESENT"
    ).count()

    return {
        "total": total,
        "present": present,
        "percentage": calculate_percentage(
            present,
            total
        ),
    }


# ==========================================
# Today's Statistics
# ==========================================

def today_statistics():

    today = date.today()

    queryset = Attendance.objects.filter(
        session__attendance_date=today
    )

    total = queryset.count()

    present = queryset.filter(
        status="PRESENT"
    ).count()

    absent = queryset.filter(
        status="ABSENT"
    ).count()

    late = queryset.filter(
        status="LATE"
    ).count()

    medical = queryset.filter(
        status="MEDICAL"
    ).count()

    return {
        "total": total,
        "present": present,
        "absent": absent,
        "late": late,
        "medical": medical,
        "percentage": calculate_percentage(
            present,
            total
        ),
    }


# ==========================================
# Monthly Statistics
# ==========================================

def monthly_statistics(year):

    result = []

    for month in range(1, 13):

        queryset = Attendance.objects.filter(
            session__attendance_date__year=year,
            session__attendance_date__month=month
        )

        total = queryset.count()

        present = queryset.filter(
            status="PRESENT"
        ).count()

        result.append({
            "month": month,
            "total": total,
            "present": present,
            "percentage": calculate_percentage(
                present,
                total
            ),
        })

    return result


# ==========================================
# Department Statistics
# ==========================================

def department_statistics():

    return Attendance.objects.values(
        "session__department__name"
    ).annotate(

        total=Count("id"),

        present=Count(
            "id",
            filter=Q(status="PRESENT")
        ),

        absent=Count(
            "id",
            filter=Q(status="ABSENT")
        )

    ).order_by(
        "session__department__name"
    )


# ==========================================
# Export CSV
# ==========================================

def export_csv(queryset):

    response = HttpResponse(
        content_type="text/csv"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="attendance_report.csv"'

    writer = csv.writer(response)

    writer.writerow([
        "Date",
        "Student",
        "Roll No",
        "Department",
        "Course",
        "Teacher",
        "Status",
        "Remarks"
    ])

    for attendance in queryset:

        writer.writerow([
            attendance.session.attendance_date,
            attendance.student.user.get_full_name(),
            attendance.student.roll_no,
            attendance.session.department.name,
            attendance.session.course.name,
            attendance.session.teacher.user.get_full_name(),
            attendance.status,
            attendance.remarks,
        ])

    return response


# ==========================================
# Export Excel
# ==========================================

def export_excel(queryset):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Attendance Report"

    headers = [
        "Date",
        "Student",
        "Roll No",
        "Department",
        "Course",
        "Teacher",
        "Status",
        "Remarks",
    ]

    for col, header in enumerate(headers, 1):

        cell = sheet.cell(row=1, column=col)

        cell.value = header

        cell.font = Font(bold=True)

    row = 2

    for attendance in queryset:

        sheet.cell(row, 1).value = str(
            attendance.session.attendance_date
        )

        sheet.cell(row, 2).value = (
            attendance.student.user.get_full_name()
        )

        sheet.cell(row, 3).value = (
            attendance.student.roll_no
        )

        sheet.cell(row, 4).value = (
            attendance.session.department.name
        )

        sheet.cell(row, 5).value = (
            attendance.session.course.name
        )

        sheet.cell(row, 6).value = (
            attendance.session.teacher.user.get_full_name()
        )

        sheet.cell(row, 7).value = (
            attendance.status
        )

        sheet.cell(row, 8).value = (
            attendance.remarks
        )

        row += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="attendance_report.xlsx"'

    workbook.save(response)

    return response


# ==========================================
# Top Students
# ==========================================

def top_students(limit=10):

    students = []

    from students.models import StudentProfile

    for student in StudentProfile.objects.select_related("user"):

        summary = student_summary(student)

        students.append({
            "student": student,
            "percentage": summary["percentage"]
        })

    students.sort(
        key=lambda x: x["percentage"],
        reverse=True
    )

    return students[:limit]


# ==========================================
# Low Attendance Students
# ==========================================

def low_attendance_students(limit=10):

    students = top_students(100000)

    students.sort(
        key=lambda x: x["percentage"]
    )

    return students[:limit]