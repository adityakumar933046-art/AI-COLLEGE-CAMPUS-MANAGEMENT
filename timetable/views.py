def clean_query_param(val):
    if val is None:
        return None
    val_str = str(val).strip()
    if val_str.lower() in ("", "none", "null", "undefined"):
        return None
    return val_str

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from datetime import datetime
from .models import Timetable

from departments.models import Department
from teachers.models import TeacherProfile
from courses.models import Course

from accounts.decorators import admin_required

from students.models import StudentProfile

from .models import Timetable
from .forms import (
    TimetableForm,
    TimetableFilterForm,
    TimetableImportForm,
)
from openpyxl import Workbook, load_workbook
import csv
import uuid
import os

# =====================================================
# TIMETABLE LIST
# =====================================================

@login_required
def timetable_list(request):

    timetables = Timetable.objects.select_related(
        "department",
        "course",
        "teacher",
        "teacher__user",
    ).filter(
        is_active=True
    )

    search = clean_query_param(request.GET.get("search"))
    if search:
        timetables = timetables.filter(
            Q(course__name__icontains=search) |
            Q(course__code__icontains=search) |
            Q(teacher__user__first_name__icontains=search) |
            Q(teacher__user__last_name__icontains=search) |
            Q(classroom__icontains=search)
        )

    department = clean_query_param(request.GET.get("department"))
    if department and department.isdigit():
        timetables = timetables.filter(department_id=int(department))

    semester = clean_query_param(request.GET.get("semester"))
    if semester and semester.isdigit():
        timetables = timetables.filter(semester=int(semester))

    section = clean_query_param(request.GET.get("section"))
    if section:
        timetables = timetables.filter(section__iexact=section)

    day = clean_query_param(request.GET.get("day"))
    if day:
        timetables = timetables.filter(day__iexact=day)

    paginator = Paginator(
        timetables.order_by(
            "day",
            "start_time"
        ),
        10
    )

    page_number = request.GET.get("page")

    page_obj = paginator.get_page(page_number)

    context = {

        "timetables": page_obj,

        "page_obj": page_obj,

        "search": search,

        "department": department,

        "semester": semester,

        "section": section,

        "day": day,

        "departments": Department.objects.all().order_by("name"),

        "filter_form": TimetableFilterForm(request.GET),

    }

    return render(
        request,
        "timetable/timetable_list.html",
        context,
    )

# =====================================================
# TIMETABLE DETAIL
# =====================================================

@login_required
def timetable_detail(request, pk):

    timetable = get_object_or_404(
        Timetable.objects.select_related(
            "department",
            "course",
            "teacher",
            "teacher__user",
        ),
        pk=pk,
    )

    context = {
        "timetable": timetable,
    }

    return render(
        request,
        "timetable/timetable_detail.html",
        context,
    )


# =====================================================
# CREATE TIMETABLE
# =====================================================

@login_required
@admin_required
def timetable_create(request):
    if request.method == "POST":
        form = TimetableForm(request.POST)
        if form.is_valid():
            dept = form.cleaned_data.get("department")
            sem = form.cleaned_data.get("semester")
            sec = form.cleaned_data.get("section")
            teacher = form.cleaned_data.get("teacher")
            day = form.cleaned_data.get("day")
            start_time = form.cleaned_data.get("start_time")
            classroom = form.cleaned_data.get("classroom")

            # Check conflicts
            if Timetable.objects.filter(teacher=teacher, day=day, start_time=start_time).exists():
                messages.error(request, f"Schedule Conflict: Faculty {teacher.full_name} is already booked on {day} at {start_time}.")
            elif Timetable.objects.filter(classroom=classroom, day=day, start_time=start_time).exists():
                messages.error(request, f"Schedule Conflict: Classroom {classroom} is already occupied on {day} at {start_time}.")
            elif Timetable.objects.filter(department=dept, semester=sem, section=sec, day=day, start_time=start_time).exists():
                messages.error(request, f"Schedule Conflict: {dept.name} Sem {sem} Sec {sec} is already scheduled on {day} at {start_time}.")
            else:
                tt_instance = form.save(commit=False)
                tt_instance.is_active = True
                tt_instance.save()
                messages.success(request, "Timetable slot created successfully.")
                return redirect("timetable_list")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = TimetableForm()

    context = {
        "form": form,
        "title": "Create Timetable",
        "button": "Save Timetable",
    }
    return render(request, "timetable/timetable_create.html", context)

# =====================================================
# STUDENT TIMETABLE
# =====================================================
@login_required
def student_timetable(request):

    student = get_object_or_404(
        StudentProfile.objects.select_related(
            "user",
            "department",
        ),
        user=request.user,
    )

    timetables = Timetable.objects.filter(
        department=student.department,
        semester=student.semester,
        is_active=True,
    ).select_related(
        "department",
        "course",
        "teacher",
        "teacher__user",
    ).order_by(
        "day",
        "start_time",
    )

    context = {
        "student": student,
        "timetables": timetables,
        "total_classes": timetables.count(),
    }

    return render(
        request,
        "timetable/student_timetable.html",
        context,
    )

# =====================================================
# TEACHER TIMETABLE
# =====================================================

@login_required
def teacher_timetable(request):

    teacher = get_object_or_404(
        TeacherProfile.objects.select_related(
            "user",
            "department",
        ),
        user=request.user,
    )

    timetables = Timetable.objects.filter(
        teacher=teacher,
        is_active=True,
    ).select_related(
        "department",
        "course",
    ).order_by(
        "day",
        "start_time",
    )
    today = datetime.now().strftime("%A").upper()

    current_time = datetime.now().time()

    today_classes = []

    for timetable in timetables:

        if timetable.day != today:
            continue

        if timetable.start_time <= current_time <= timetable.end_time:
            timetable.session_status = "ONGOING"

        elif current_time < timetable.start_time:
            timetable.session_status = "UPCOMING"

        else:
            timetable.session_status = "COMPLETED"

        today_classes.append(timetable)

    context = {

    "teacher": teacher,

    "today_classes": today_classes,

    "timetables": timetables,

    "total_classes": timetables.count(),

}

    return render(
        request,
        "timetable/teacher_timetable.html",
        context,
    )
# =====================================================
# TIMETABLE DASHBOARD
# =====================================================

@login_required
def timetable_dashboard(request):

    total_timetables = Timetable.objects.count()

    active_timetables = Timetable.objects.filter(
        is_active=True
    ).count()

    departments = Timetable.objects.values(
        "department__name"
    ).distinct().count()

    teachers = Timetable.objects.values(
        "teacher"
    ).distinct().count()

    recent_timetables = Timetable.objects.select_related(
        "department",
        "course",
        "teacher",
        "teacher__user",
    ).order_by(
        "-created_at"
    )[:10]

    context = {

        "total_timetables": total_timetables,

        "active_timetables": active_timetables,

        "departments": departments,

        "teachers": teachers,

        "recent_timetables": recent_timetables,

    }

    return render(
        request,
        "timetable/dashboard.html",
        context,
    )


# =====================================================
# EXPORT CSV
# =====================================================

@login_required
def export_timetable_csv(request):

    response = HttpResponse(
        content_type="text/csv"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="timetable.csv"'

    writer = csv.writer(response)

    writer.writerow([
        "Department",
        "Semester",
        "Section",
        "Course",
        "Teacher",
        "Day",
        "Start Time",
        "End Time",
        "Classroom",
    ])

    timetables = Timetable.objects.select_related(
        "department",
        "course",
        "teacher",
        "teacher__user",
    )

    for timetable in timetables:

        writer.writerow([
            timetable.department.name,
            timetable.semester,
            timetable.section,
            timetable.course.name,
            timetable.teacher.user.get_full_name(),
            timetable.day,
            timetable.start_time,
            timetable.end_time,
            timetable.classroom,
        ])

    return response


# =====================================================
# UPDATE TIMETABLE
# =====================================================

@login_required
@admin_required
def timetable_update(request, pk):

    timetable = get_object_or_404(Timetable, pk=pk)

    if request.method == "POST":

        form = TimetableForm(request.POST, instance=timetable)

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Timetable updated successfully."
            )

            return redirect("timetable_detail", pk=timetable.pk)

    else:

        form = TimetableForm(instance=timetable)

    context = {
        "form": form,
        "timetable": timetable,
    }

    return render(
        request,
        "timetable/timetable_update.html",
        context,
    )


# =====================================================
# DELETE TIMETABLE
# =====================================================

@login_required
@admin_required
def timetable_delete(request, pk):

    timetable = get_object_or_404(Timetable, pk=pk)

    if request.method == "POST":

        timetable.delete()

        messages.success(
            request,
            "Timetable deleted successfully."
        )

        return redirect("timetable_list")

    return render(
        request,
        "timetable/timetable_delete.html",
        {
            "timetable": timetable,
        },
    )


# =====================================================
# EXPORT EXCEL
# =====================================================

@login_required
def export_timetable_excel(request):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Timetable"

    sheet.append([
        "Department",
        "Semester",
        "Section",
        "Course",
        "Teacher",
        "Day",
        "Start Time",
        "End Time",
        "Classroom",
    ])

    timetables = Timetable.objects.select_related(
        "department",
        "course",
        "teacher",
        "teacher__user",
    )

    for timetable in timetables:

        sheet.append([
            timetable.department.name,
            timetable.semester,
            timetable.section,
            timetable.course.name,
            timetable.teacher.user.get_full_name(),
            timetable.day,
            str(timetable.start_time),
            str(timetable.end_time),
            timetable.classroom,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="timetable.xlsx"'

    workbook.save(response)

    return response




# =====================================================
# IMPORT TIMETABLE (PART 1)
# =====================================================

@login_required
@admin_required
def import_timetable(request):

    if request.method == "POST":

        form = TimetableImportForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():

            file = request.FILES["timetable_file"]

            imported = 0
            skipped = 0
            errors = []

            extension = os.path.splitext(
                file.name
            )[1].lower()

            rows = []

            # -------------------------------
            # EXCEL FILE
            # -------------------------------

            if extension in [".xlsx", ".xls"]:

                workbook = load_workbook(file)

                sheet = workbook.active

                for row in sheet.iter_rows(
                    min_row=2,
                    values_only=True,
                ):

                    rows.append(row)

            # -------------------------------
            # CSV FILE
            # -------------------------------

            elif extension == ".csv":

                decoded = file.read().decode(
                    "utf-8"
                ).splitlines()

                reader = csv.reader(decoded)

                next(reader)

                for row in reader:

                    rows.append(row)

            else:

                messages.error(
                    request,
                    "Unsupported file format."
                )

                return redirect(
                    "import_timetable"
                )

            # --------------------------------
            # LOOP STARTS HERE
            # --------------------------------

            for index, row in enumerate(
                rows,
                start=2
            ):

                try:

                    if not row:

                        continue

                    department_name = str(
                        row[0]
                    ).strip()

                    semester = int(row[1])

                    section = str(
                        row[2]
                    ).strip()

                    course_name = str(
                        row[3]
                    ).strip()

                    teacher_name = str(
                        row[4]
                    ).strip()

                    day = str(
                        row[5]
                    ).strip().upper()

                    start_time = row[6]

                    end_time = row[7]

                    classroom = str(
                        row[8]
                    ).strip()


                                        # -------------------------
                    # FETCH OBJECTS
                    # -------------------------

                    department = Department.objects.get(
                        name__iexact=department_name
                    )

                    course = Course.objects.get(
                        name__iexact=course_name
                    )

                    teacher = TeacherProfile.objects.get(
                        user__first_name__iexact=teacher_name.split()[0]
                    )

                    # -------------------------
                    # DUPLICATE CHECK
                    # -------------------------

                    if Timetable.objects.filter(

                        department=department,

                        semester=semester,

                        section=section,

                        day=day,

                        start_time=start_time,

                    ).exists():

                        skipped += 1

                        continue

                    # -------------------------
                    # TEACHER CONFLICT
                    # -------------------------

                    if Timetable.objects.filter(

                        teacher=teacher,

                        day=day,

                        start_time=start_time,

                    ).exists():

                        errors.append(

                            f"Row {index}: Teacher already has another class."

                        )

                        continue

                    # -------------------------
                    # CLASSROOM CONFLICT
                    # -------------------------

                    if Timetable.objects.filter(

                        classroom=classroom,

                        day=day,

                        start_time=start_time,

                    ).exists():

                        errors.append(

                            f"Row {index}: Classroom already occupied."

                        )

                        continue

                    # -------------------------
                    # CREATE TIMETABLE
                    # -------------------------

                    Timetable.objects.create(

                        department=department,

                        semester=semester,

                        section=section,

                        course=course,

                        teacher=teacher,

                        day=day,

                        start_time=start_time,

                        end_time=end_time,

                        classroom=classroom,

                        is_active=True,

                    )

                    imported += 1

                except Exception as e:

                    errors.append(

                        f"Row {index}: {str(e)}"

                    )

            # -------------------------
            # SUCCESS MESSAGE
            # -------------------------

            messages.success(

                request,

                f"{imported} timetable(s) imported successfully."

            )

            if skipped:

                messages.warning(

                    request,

                    f"{skipped} duplicate record(s) skipped."

                )

            if errors:

                for error in errors:

                    messages.error(

                        request,

                        error

                    )

            return redirect(
                "timetable_list"
            )

    else:

        form = TimetableImportForm()

    return render(

        request,

        "timetable/import_timetable.html",

        {

            "form": form,

        },

    )


# =====================================================
# DOWNLOAD TIMETABLE SAMPLE
# =====================================================

@login_required
@admin_required
def download_timetable_sample(request):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Sample Timetable"

    # Header
    sheet.append([
        "Department",
        "Semester",
        "Section",
        "Course",
        "Teacher",
        "Day",
        "Start Time",
        "End Time",
        "Classroom",
    ])

    # Sample Row 1
    sheet.append([
        "Computer Science",
        5,
        "A",
        "Database Management System",
        "Rahul",
        "MONDAY",
        "09:00",
        "10:00",
        "A-101",
    ])

    # Sample Row 2
    sheet.append([
        "Computer Science",
        5,
        "A",
        "Python Programming",
        "Rahul",
        "TUESDAY",
        "10:00",
        "11:00",
        "A-102",
    ])

    # Sample Row 3
    sheet.append([
        "Information Technology",
        3,
        "B",
        "Operating System",
        "Amit",
        "WEDNESDAY",
        "11:00",
        "12:00",
        "B-201",
    ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="sample_timetable.xlsx"'

    workbook.save(response)

    return response
