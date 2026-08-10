from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils.crypto import get_random_string


from .models import StudentProfile
from io import BytesIO
from openpyxl import Workbook


User = get_user_model()

def generate_student_username(first_name, roll_no="", admission_no=""):
    base_name = (first_name or "student").strip().lower().replace(" ", "")
    clean_roll = str(roll_no or admission_no or "").strip()
    if clean_roll:
        base_username = f"{base_name}{clean_roll}"
    else:
        base_username = base_name

    username = base_username
    count = 1
    while User.objects.filter(username=username).exists():
        username = f"{base_username}{count}"
        count += 1
    return username




# ==========================================================
# CREATE STUDENT
# ==========================================================

@transaction.atomic
def create_student(data, created_by=None):


    try:

        # ==========================
        # CREATE USER
        # ==========================

        raw_username = (data.get("username") or "").strip()
        first_name = (data.get("first_name") or "").strip()
        last_name = (data.get("last_name") or "").strip()
        email = (data.get("email") or "").strip().lower()
        roll_no = (data.get("roll_no") or "").strip()
        admission_no = (data.get("admission_no") or "").strip()

        if not raw_username or User.objects.filter(username=raw_username).exists():
            username = generate_student_username(first_name, roll_no, admission_no)
        else:
            username = raw_username

        password = data.get("password")
        if not password:
            password = get_random_string(10)

        user = User.objects.create_user(
            username=username,
            first_name=first_name,
            last_name=last_name,
            email=email,
            password=password,
            role="STUDENT",
        )
        user.must_change_password = True
        user.save()



        # ==========================
        # CREATE STUDENT PROFILE
        # ==========================

        student = StudentProfile.objects.create(

            user=user,


            admission_no=data.get(
                "admission_no"
            ),


            roll_no=data.get(
                "roll_no"
            ),


            department=data.get(
                "department"
            ),


            semester=data.get(
                "semester"
            ),


            section=data.get(
                "section"
            ),


            batch=data.get(
                "batch"
            ),


            academic_year=data.get(
                "academic_year"
            ),


            admission_date=data.get(
                "admission_date"
            ),



            photo=data.get(
                "photo"
            ),



            phone=data.get(
                "phone"
            ),


            gender=data.get(
                "gender"
            ),


            date_of_birth=data.get(
                "date_of_birth"
            ),


            blood_group=data.get(
                "blood_group"
            ),



            father_name=data.get(
                "father_name"
            ),


            mother_name=data.get(
                "mother_name"
            ),


            guardian_phone=data.get(
                "guardian_phone"
            ),



            address=data.get(
                "address"
            ),


            city=data.get(
                "city"
            ),


            state=data.get(
                "state"
            ),


            pincode=data.get(
                "pincode"
            ),



            cgpa=data.get("cgpa") if data.get("cgpa") is not None else 0.00,

            status=data.get("status") or "ACTIVE",


            created_by=created_by,

        )



        return {

            "success": True,

            "student": student,

            "user": user,

            "username": user.username,

            "password": password,

            "errors": []

        }



    except Exception as e:


        return {

            "success": False,

            "student": None,

            "username": None,

            "password": None,

            "errors": [
                str(e)
            ]

        }
    # ==========================================================
# UPDATE STUDENT
# ==========================================================

@transaction.atomic
def update_student(student, data, updated_by=None):


    try:

        user = student.user


        # ==========================
        # UPDATE USER INFORMATION
        # ==========================

        if data.get("username"):

            user.username = data.get(
                "username"
            )


        if data.get("first_name"):

            user.first_name = data.get(
                "first_name"
            )


        if data.get("last_name"):

            user.last_name = data.get(
                "last_name"
            )


        if data.get("email"):

            user.email = data.get(
                "email"
            )


        if data.get("password"):

            user.set_password(
                data.get("password")
            )


        user.save()



        # ==========================
        # UPDATE STUDENT PROFILE
        # ==========================

        update_fields = [

            "admission_no",
            "roll_no",
            "department",
            "semester",
            "section",
            "batch",
            "academic_year",
            "admission_date",

            "photo",
            "phone",
            "gender",
            "date_of_birth",
            "blood_group",

            "father_name",
            "mother_name",
            "guardian_phone",

            "address",
            "city",
            "state",
            "pincode",

            "cgpa",
            "status",

        ]



        for field in update_fields:


            if field in data:

                setattr(
                    student,
                    field,
                    data.get(field)
                )



        student.updated_by = updated_by


        student.save()



        return {

            "success": True,

            "student": student,

            "errors": []

        }



    except Exception as e:


        return {

            "success": False,

            "student": None,

            "errors": [
                str(e)
            ]

        }





# ==========================================================
# DELETE STUDENT
# ==========================================================

@transaction.atomic
def delete_student(student):


    try:

        user = student.user


        if user:

            user.delete()

        else:

            student.delete()



        return {

            "success": True,

            "errors": []

        }



    except Exception as e:


        return {

            "success": False,

            "errors": [
                str(e)
            ]

        }





# ==========================================================
# ACTIVATE STUDENT
# ==========================================================

def activate_student(student, user=None):


    student.status = "ACTIVE"


    if user:

        student.updated_by = user


    student.save()



    return student





# ==========================================================
# DEACTIVATE STUDENT
# ==========================================================

def deactivate_student(student, user=None):


    student.status = "INACTIVE"


    if user:

        student.updated_by = user


    student.save()



    return student

import pandas as pd

from datetime import datetime



# ==========================================================
# BULK IMPORT STUDENTS FROM EXCEL
# ==========================================================

@transaction.atomic
def bulk_import_students(records, import_batch=None):


    success_count = 0

    failed_count = 0


    credentials = []

    errors = []



    for index, row in enumerate(records, start=2):


        try:


            username = row.get(
                "username"
            )


            password = get_random_string(
                8
            )



            # ==========================
            # CREATE USER
            # ==========================


            user = User.objects.create_user(

                username=username,

                first_name=row.get(
                    "first_name",
                    ""
                ),

                last_name=row.get(
                    "last_name",
                    ""
                ),

                email=row.get(
                    "email",
                    ""
                ),

                password=password,

            )



            # ==========================
            # CREATE STUDENT
            # ==========================


            student = StudentProfile.objects.create(

                user=user,


                admission_no=row.get(
                    "admission_no"
                ),


                roll_no=row.get(
                    "roll_no"
                ),


                department_id=row.get(
                    "department"
                ),


                semester=row.get(
                    "semester",
                    1
                ),


                section=row.get(
                    "section",
                    ""
                ),


                batch=row.get(
                    "batch",
                    ""
                ),


                academic_year=row.get(
                    "academic_year",
                    ""
                ),


                phone=row.get(
                    "phone",
                    ""
                ),


                father_name=row.get(
                    "father_name",
                    ""
                ),


                mother_name=row.get(
                    "mother_name",
                    ""
                ),


                guardian_phone=row.get(
                    "guardian_phone",
                    ""
                ),


                address=row.get(
                    "address",
                    ""
                ),


                cgpa=row.get(
                    "cgpa",
                    0
                ),


                status="ACTIVE",


                is_imported=True,


                import_batch=import_batch,

            )



            credentials.append({

                "Username": username,

                "Password": password,

                "Student Name": user.get_full_name(),

                "Roll No": student.roll_no,

                "Admission No": student.admission_no,

            })



            success_count += 1



        except Exception as e:


            failed_count += 1


            errors.append({

                "row": index,

                "error": str(e),

                "data": dict(row),

            })



    return {


        "total_records": len(records),


        "success_count": success_count,


        "failed_count": failed_count,


        "credentials": credentials,


        "errors": errors,


    }





# ==========================================================
# IMPORT SUMMARY
# ==========================================================

def get_import_summary(result):


    return {


        "total_records": result.get(
            "total_records",
            0
        ),


        "success_count": result.get(
            "success_count",
            0
        ),


        "failed_count": result.get(
            "failed_count",
            0
        ),

    }
# ==========================================================
# GENERATE STUDENT CREDENTIALS EXCEL
# ==========================================================

def generate_credentials_excel(credentials):

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Student Credentials"


    worksheet.append([

        "Username",
        "Password",
        "Student Name",
        "Roll No",
        "Admission No",

    ])


    for data in credentials:

        worksheet.append([

            data.get("Username"),

            data.get("Password"),

            data.get("Student Name"),

            data.get("Roll No"),

            data.get("Admission No"),

        ])



    output = BytesIO()

    workbook.save(output)


    output.seek(0)


    return output
# ==========================================================
# GENERATE IMPORT ERROR REPORT
# ==========================================================

def generate_error_report(errors):


    workbook = Workbook()


    worksheet = workbook.active

    worksheet.title = "Import Errors"



    worksheet.append([

        "Row Number",
        "Error",
        "Data",

    ])



    for error in errors:


        worksheet.append([

            error.get("row"),

            error.get("error"),

            str(
                error.get("data")
            ),

        ])



    output = BytesIO()


    workbook.save(output)


    output.seek(0)


    return output

# ==========================================================
# STUDENT SAMPLE EXCEL
# ==========================================================

def generate_student_sample_excel():


    workbook = Workbook()


    worksheet = workbook.active

    worksheet.title = "Students"



    worksheet.append([

        "username",
        "first_name",
        "last_name",
        "email",

        "admission_no",
        "roll_no",

        "department",
        "semester",

        "section",
        "batch",

        "academic_year",

        "phone",

        "father_name",
        "mother_name",

        "guardian_phone",

        "cgpa",

    ])



    worksheet.append([

        "student001",
        "Rahul",
        "Sharma",
        "rahul@gmail.com",

        "ADM001",
        "CSE001",

        "CSE",
        3,

        "A",
        "2025",

        "2025-26",

        "9876543210",

        "Father Name",
        "Mother Name",

        "9876543211",

        8.5,

    ])



    output = BytesIO()


    workbook.save(output)


    output.seek(0)


    return output

# ==========================================================
# EXPORT STUDENTS EXCEL
# ==========================================================

def export_students_excel():


    workbook = Workbook()


    worksheet = workbook.active

    worksheet.title = "Students"



    worksheet.append([

        "Username",
        "Name",
        "Email",

        "Admission No",
        "Roll No",

        "Department",

        "Semester",

        "Phone",

        "Status",

    ])



    students = StudentProfile.objects.select_related(
        "user",
        "department"
    )



    for student in students:


        worksheet.append([


            student.user.username,


            student.full_name,


            student.user.email,


            student.admission_no,


            student.roll_no,


            student.department.name
            if student.department else "",


            student.semester,


            student.phone,


            student.status,


        ])



    output = BytesIO()


    workbook.save(output)


    output.seek(0)


    return output
# ==========================================================
# EXPORT STUDENTS CSV
# ==========================================================

def export_students_csv():


    import csv


    response = HttpResponse(
        content_type="text/csv"
    )


    writer = csv.writer(response)



    writer.writerow([

        "Username",
        "Name",
        "Email",
        "Admission No",
        "Roll No",
        "Department",
        "Semester",
        "Phone",
        "Status",

    ])



    students = StudentProfile.objects.select_related(
        "user",
        "department"
    )



    for student in students:


        writer.writerow([

            student.user.username,

            student.full_name,

            student.user.email,

            student.admission_no,

            student.roll_no,

            student.department.name
            if student.department else "",

            student.semester,

            student.phone,

            student.status,

        ])



    return response